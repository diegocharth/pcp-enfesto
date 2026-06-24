"""A3: o cabecalho das planilhas deve conter todos os parametros do calculo."""
import tempfile
import openpyxl
from engine.alocador_rolos import alocar_rolos
from exportar.export_xlsx import (
    _resumo_parametros_txt,
    _aba_resumo_multiref,
    exportar_alocacao,
)


def _cfg():
    return {
        "consumo_peca_m": 1.0645,
        "mesa_comprimento_m": 10,
        "limite_folhas_padrao": 70,
        "desvio_absoluto_padrao": 4,
        "desvio_percentual_padrao": 20,
        "criterio_combinacao": "MIN",
        "num_opcoes_saida": 2,
        "timeout": 120,
        "tempo_processamento_s": 12.3,
        "versao": "2.10.1",
        "regras_especiais": {"G": {"hi": 0}},
    }


def test_resumo_parametros_txt_tem_todos_os_campos():
    s = _resumo_parametros_txt(_cfg())
    for marca in ["Consumo", "Mesa", "Folhas", "Tol. abs", "Tol. %",
                  "Criterio", "Opcoes", "Timeout", "Tempo real",
                  "Limites especiais", "Versao"]:
        assert marca in s, f"faltou '{marca}' em: {s}"
    assert "1.0645" in s
    assert "MIN" in s
    assert "120" in s
    assert "2.10.1" in s
    assert "G" in s  # limite especial


def test_resumo_parametros_txt_sem_campos_opcionais():
    """Sem timeout/tempo/regras nao deve quebrar."""
    s = _resumo_parametros_txt({"consumo_peca_m": 1.0, "mesa_comprimento_m": 10,
                                "limite_folhas_padrao": 70})
    assert "Consumo" in s
    assert "Timeout" in s          # mostra "—"
    assert "Limites especiais" in s  # mostra "nenhum"


def test_aba_resumo_multiref_cabecalho_tem_params():
    sol = {
        "n_mapas": 1,
        "refs_sol": [{"nome": "REF1", "mapas": []}],
        "comprimentos": [],
        "resumo": {"total_folhas": 10, "desvio_total": 2,
                   "comprimento_total": 50.0, "media_pecas_mapa": 6.0},
    }
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _aba_resumo_multiref(wb, [sol], "GrupoX", _cfg())
    a2 = wb["Resumo"]["A2"].value
    assert "Tol. abs" in a2 and "Criterio" in a2 and "Versao" in a2


def test_export_alocacao_mostra_parametros():
    plano = {"mapas": [{"id": 0, "n_pecas": 4}],
             "camadas": {"AZUL": {0: 3}}, "consumo_peca": 1.0}
    cfg = {"margem_seguranca_enfesto_m": 0.10, "folga_incerteza_pct": 0.03,
           "folga_incerteza_m": 0.0, "ponta_minima_util_m": 0.5}
    res = alocar_rolos(plano, {"AZUL": [20.0]}, cfg)

    with tempfile.TemporaryDirectory() as d:
        params = {**res["params"], "versao": "2.10.1"}
        cam = exportar_alocacao(res, "TESTE", d, params)
        wb = openpyxl.load_workbook(cam)
        ws = wb["Resumo Alocacao"]
        textos = " ".join(str(c.value) for row in ws.iter_rows()
                           for c in row if c.value is not None)
    assert "Margem" in textos
    assert "Folga" in textos
    assert "Ponta minima" in textos
    assert "2.10.1" in textos
