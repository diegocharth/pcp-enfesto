"""
PCP Enfestos — Exportador Excel v2.2
Abas: Resumo + Trade-off | Por Opção: Plano + Conferência | Comparativo
"""

import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

# ── Paleta ────────────────────────────────────────────────────────────────────
C_AZUL_ESC  = "1A3A5C"
C_AZUL_MED  = "2E6DA4"
C_AZUL_CLR  = "D6E4F0"
C_CINZA_HDR = "F0EEE8"
C_CINZA_ALT = "F8F7F4"
C_VERDE     = "D5F5E3"
C_VERDE_TX  = "1D6A2E"
C_LARANJA   = "FDEBD0"
C_LARANJA_TX= "7D4C00"
C_VERMELHO  = "FADBD8"
C_VERM_TX   = "922B21"
C_AMARELO   = "FFFBEA"
C_AMAR_TX   = "5A4000"
C_BRANCO    = "FFFFFF"
C_TOTAL     = "E8E6DF"
C_TOTAL_TX  = "1A3A5C"
C_DESVIO0   = "D5F5E3"  # desvio zero = verde
C_DESVIO_OK = "EBF5FB"  # desvio ok = azul claro
C_DESVIO_ER = "FADBD8"  # desvio fora = vermelho


def _bd(cor="CCCCCC"):
    s = Side(style="thin", color=cor)
    return Border(left=s, right=s, top=s, bottom=s)


def _cel(ws, row, col, val=None, negrito=False, alinha="left",
         fundo=None, cor_txt=None, tamanho=10, wrap=False, numero=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Calibri", size=tamanho, bold=negrito,
                       color=cor_txt or "000000")
    c.alignment = Alignment(horizontal=alinha, vertical="center",
                            wrap_text=wrap)
    c.border    = _bd()
    if fundo:
        c.fill = PatternFill("solid", fgColor=fundo)
    return c


def _cabecalho_sheet(ws, titulo, subtitulo=""):
    ws.row_dimensions[1].height = 26
    ws.merge_cells(f"A1:{get_column_letter(ws.max_column or 20)}1")
    c = ws["A1"]
    c.value     = titulo
    c.font      = Font(name="Calibri", size=14, bold=True, color=C_BRANCO)
    c.fill      = PatternFill("solid", fgColor=C_AZUL_ESC)
    c.alignment = Alignment(horizontal="left", vertical="center")
    if subtitulo:
        ws.row_dimensions[2].height = 18
        ws.merge_cells(f"A2:{get_column_letter(ws.max_column or 20)}2")
        c2 = ws["A2"]
        c2.value     = subtitulo
        c2.font      = Font(name="Calibri", size=9, color="666666")
        c2.fill      = PatternFill("solid", fgColor=C_AZUL_CLR)
        c2.alignment = Alignment(horizontal="left", vertical="center")
        return 3
    return 2


def _linha_titulo_secao(ws, row, texto, ncols):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1, value=texto)
    c.font      = Font(name="Calibri", size=10, bold=True, color=C_BRANCO)
    c.fill      = PatternFill("solid", fgColor=C_AZUL_MED)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = _bd()
    ws.row_dimensions[row].height = 18
    return row + 1


def _linha_vazia(ws, row):
    ws.row_dimensions[row].height = 8
    return row + 1


# ── Funções de análise de trade-off ──────────────────────────────────────────
def _calcular_tradeoff(sols, grade, tamanhos):
    if len(sols) < 2:
        return None
    s0, s1 = sols[0]["resumo"], sols[1]["resumo"]
    d0, d1 = s0.get("desvio_total", 0), s1.get("desvio_total", 0)
    p0, p1 = s0.get("media_pecas_mapa", 0), s1.get("media_pecas_mapa", 0)
    diff_dev = d1 - d0
    diff_ppm = p1 - p0
    pct_ppm  = round((diff_ppm / p0 * 100), 1) if p0 > 0 else 0

    if d0 == 0 and d1 == 0:
        msg = f"Ambas as opções sem ajuste de grade. Prefira Opção {1 if p0 >= p1 else 2} (maior eficiência de encaixe)."
        nivel = "ok"
    elif d0 == 0:
        msg = f"Opção 1 sem ajuste de grade. Opção 2 ajusta {d1} peças" + (f" e ganha +{pct_ppm}% de peças/mapa." if pct_ppm > 0 else " sem ganho adicional de encaixe.")
        nivel = "ok"
    elif d1 == 0:
        msg = f"Opção 2 sem ajuste de grade. Opção 1 ajusta {d0} peças" + (f" com +{abs(pct_ppm)}% mais peças/mapa." if pct_ppm < 0 else ".")
        nivel = "ok"
    elif diff_dev > 0 and diff_ppm > 0:
        justifica = pct_ppm >= 10
        msg = f"Opção 2 precisa de +{diff_dev} peças de ajuste mas ganha +{pct_ppm}% de eficiência de encaixe. {'Ganho significativo — pode justificar o ajuste.' if justifica else 'Ganho pequeno — prefira Opção 1 se o ajuste for crítico.'}"
        nivel = "ok" if justifica else "aviso"
    elif diff_dev <= 0:
        msg = f"Opção 2 tem menor ajuste ({d1} vs {d0} peças)." + (" Recomenda-se Opção 2." if diff_ppm >= 0 else f" Mas Opção 1 tem {abs(pct_ppm)}% mais peças/mapa.")
        nivel = "ok"
    else:
        msg = f"Opção 1: {d0} peças de ajuste, {p0} pç/mapa. Opção 2: {d1} peças de ajuste, {p1} pç/mapa."
        nivel = "info"

    return {"msg": msg, "nivel": nivel, "d0": d0, "d1": d1, "p0": p0, "p1": p1,
            "diff_dev": diff_dev, "pct_ppm": pct_ppm}


# ══════════════════════════════════════════════════════════════════════════════
# ABA RESUMO
# ══════════════════════════════════════════════════════════════════════════════
def _resumo_parametros_txt(config):
    """Monta a linha de parametros do calculo para o cabecalho das planilhas.
    Le tudo de `config` (fonte unica). Campos ausentes viram '—'/'nenhum'."""
    def g(k, d=None):
        v = config.get(k, d)
        return d if v is None else v

    regras = config.get("regras_especiais") or {}
    if regras:
        partes = []
        for t, r in regras.items():
            lo = r.get("lo", "")
            hi = r.get("hi", "")
            partes.append(f"{t}[{lo}..{hi}]")
        regras_txt = ", ".join(partes)
    else:
        regras_txt = "nenhum"

    timeout = config.get("timeout")
    tempo   = config.get("tempo_processamento_s")
    return (
        f"Consumo: {g('consumo_peca_m', 1.0645)} m/pc"
        f"  |  Mesa: {g('mesa_comprimento_m', 10)} m"
        f"  |  Folhas/enfesto: {g('limite_folhas_padrao', 70)}"
        f"  |  Tol. abs: {g('desvio_absoluto_padrao', '—')} pc"
        f"  |  Tol. %: {g('desvio_percentual_padrao', '—')}%"
        f"  |  Criterio: {g('criterio_combinacao', '—')}"
        f"  |  Opcoes: {g('num_opcoes_saida', '—')}"
        f"  |  Timeout: {timeout if timeout is not None else '—'} s"
        f"  |  Tempo real: {round(tempo, 1) if tempo is not None else '—'} s"
        f"  |  Limites especiais: {regras_txt}"
        f"  |  Versao {g('versao', '—')}"
    )


def _aba_resumo(wb, solucoes, grade, tamanhos, referencia, config):
    ws = wb.create_sheet("Resumo")
    ws.column_dimensions["A"].width = 32
    for i in range(2, 2 + len(solucoes)):
        ws.column_dimensions[get_column_letter(i)].width = 22

    ts = datetime.now().strftime("%d/%m/%Y %H:%M")
    n  = len(solucoes)
    ncols = max(n + 1, 3)

    # Cabeçalho
    ws.merge_cells(f"A1:{get_column_letter(ncols + 1)}1")
    c = ws["A1"]
    c.value = f"PCP ENFESTOS — {referencia.upper()}"
    c.font  = Font(name="Calibri", size=14, bold=True, color=C_BRANCO)
    c.fill  = PatternFill("solid", fgColor=C_AZUL_ESC)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{get_column_letter(ncols + 1)}2")
    c2 = ws["A2"]
    c2.value = f"Gerado em {ts}  |  " + _resumo_parametros_txt(config)
    c2.font  = Font(name="Calibri", size=9, color="444444")
    c2.fill  = PatternFill("solid", fgColor=C_AZUL_CLR)
    c2.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    row = 4
    row = _linha_titulo_secao(ws, row, "COMPARATIVO DE OPÇÕES", ncols + 1)

    # Headers das opções
    _cel(ws, row, 1, "Métrica", negrito=True, fundo=C_CINZA_HDR, tamanho=10)
    for i, s in enumerate(solucoes, 2):
        _cel(ws, row, i, f"Opção {i-1}", negrito=True, fundo=C_AZUL_MED,
             cor_txt=C_BRANCO, alinha="center", tamanho=10)
    ws.row_dimensions[row].height = 18
    row += 1

    def linha_metrica(label, vals, fundo_lin=None):
        nonlocal row
        _cel(ws, row, 1, label, fundo=fundo_lin or C_CINZA_ALT)
        for i, v in enumerate(vals, 2):
            _cel(ws, row, i, v, alinha="center", fundo=fundo_lin or C_BRANCO)
        ws.row_dimensions[row].height = 18
        row += 1

    metricas = [
        ("Nº de enfestos",           [s["resumo"]["n_mapas"] for s in solucoes]),
        ("Total de folhas",           [s["resumo"]["total_folhas"] for s in solucoes]),
        ("Desvio de grade (peças)",   [s["resumo"].get("desvio_total", "—") for s in solucoes]),
        ("Média de peças por mapa",   [s["resumo"]["media_pecas_mapa"] for s in solucoes]),
        ("Comprimento total (m)",     [s["resumo"]["comprimento_total"] for s in solucoes]),
    ]
    for label, vals in metricas:
        linha_metrica(label, vals)

    # Linha de ajuste de grade com cor
    row_dev = row
    _cel(ws, row, 1, "Ajuste de grade", negrito=True, fundo=C_CINZA_HDR)
    for i, s in enumerate(solucoes, 2):
        dev = s["resumo"].get("desvio_total", 0)
        if dev == 0:
            _cel(ws, row, i, "✓ Sem ajuste", negrito=True, alinha="center",
                 fundo=C_DESVIO0, cor_txt=C_VERDE_TX)
        elif dev <= 15:
            _cel(ws, row, i, f"{dev} peças", alinha="center",
                 fundo=C_AMARELO, cor_txt=C_AMAR_TX)
        else:
            _cel(ws, row, i, f"{dev} peças", alinha="center",
                 fundo=C_VERMELHO, cor_txt=C_VERM_TX)
    ws.row_dimensions[row].height = 18
    row += 1

    row = _linha_vazia(ws, row)

    # Análise trade-off
    if len(solucoes) >= 2:
        tf = _calcular_tradeoff(solucoes, grade, tamanhos)
        if tf:
            row = _linha_titulo_secao(ws, row, "ANÁLISE DE TRADE-OFF", ncols + 1)
            fundo_tf = {
                "ok":    C_VERDE,
                "aviso": C_AMARELO,
                "info":  C_AZUL_CLR
            }.get(tf["nivel"], C_BRANCO)
            cor_tf = {
                "ok":    C_VERDE_TX,
                "aviso": C_AMAR_TX,
                "info":  C_AZUL_ESC
            }.get(tf["nivel"], "000000")

            ws.merge_cells(f"A{row}:{get_column_letter(ncols + 1)}{row}")
            c = ws.cell(row=row, column=1, value=tf["msg"])
            c.font      = Font(name="Calibri", size=10, color=cor_tf)
            c.fill      = PatternFill("solid", fgColor=fundo_tf)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border    = _bd()
            ws.row_dimensions[row].height = 32
            row += 1

            # Tabela comparativa rápida
            row = _linha_vazia(ws, row)
            for label, v0, v1 in [
                ("Opção 1 — Desvio", tf["d0"], None),
                ("Opção 1 — Média peças/mapa", tf["p0"], None),
                ("Opção 2 — Desvio", tf["d1"], None),
                ("Opção 2 — Média peças/mapa", tf["p1"], None),
            ]:
                _cel(ws, row, 1, label, fundo=C_CINZA_ALT)
                _cel(ws, row, 2, v0, alinha="center", fundo=C_BRANCO)
                ws.row_dimensions[row].height = 16
                row += 1

    row = _linha_vazia(ws, row)

    # Resumo dos mapas por opção
    row = _linha_titulo_secao(ws, row, "ESTRUTURA DOS ENFESTOS POR OPÇÃO", ncols + 1)
    for i, sol in enumerate(solucoes, 1):
        _cel(ws, row, 1, f"Opção {i}:", negrito=True, fundo=C_CINZA_HDR)
        ws.row_dimensions[row].height = 16
        row += 1
        for mi, mapa in enumerate(sol["mapas"]):
            comp = sol["resumo"]["comprimento_por_mapa"][mi] if "comprimento_por_mapa" in sol["resumo"] else "—"
            ms = "+".join(f"{mapa[t]}{t}" for t in tamanhos if mapa.get(t, 0) > 0)
            npecas = sum(mapa.values())
            tf_enf = sum(sol["folhas"][c][mi] for c in grade)
            linha = f"  Enfesto {mi+1}: {ms} = {npecas} peças × {config.get('consumo_peca_m',1.0645)}m = {comp}m  |  {tf_enf} folhas"
            ws.merge_cells(f"A{row}:{get_column_letter(ncols + 1)}{row}")
            c = ws.cell(row=row, column=1, value=linha)
            c.font      = Font(name="Calibri", size=9)
            c.fill      = PatternFill("solid", fgColor=C_BRANCO if mi % 2 == 0 else C_CINZA_ALT)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border    = _bd()
            ws.row_dimensions[row].height = 16
            row += 1
        row = _linha_vazia(ws, row)

    return ws


# ══════════════════════════════════════════════════════════════════════════════
# ABA PLANO POR OPÇÃO
# ══════════════════════════════════════════════════════════════════════════════
def _aba_plano(wb, sol, idx, grade, tamanhos, referencia, config):
    ws = wb.create_sheet(f"Op{idx}_Plano")
    cores = list(grade.keys())
    consumo = float(config.get("consumo_peca_m", 1.0645))
    r = sol["resumo"]

    # Larguras
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 10
    for i in range(3, 3 + len(tamanhos)):
        ws.column_dimensions[get_column_letter(i)].width = 9
    ws.column_dimensions[get_column_letter(3 + len(tamanhos))].width = 10

    # Cabeçalho
    ncols = 2 + len(tamanhos) + 1
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value = f"PLANO DE CORTE — {referencia.upper()} — OPÇÃO {idx}"
    c.font  = Font(name="Calibri", size=13, bold=True, color=C_BRANCO)
    c.fill  = PatternFill("solid", fgColor=C_AZUL_ESC)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # KPIs
    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    dev = r.get("desvio_total", 0)
    kpi = f"Enfestos: {r['n_mapas']}  |  Total folhas: {r['total_folhas']}  |  Ajuste de grade: {'Sem ajuste ✓' if dev == 0 else str(dev)+' peças'}  |  Média peças/mapa: {r['media_pecas_mapa']}  |  Comp. total: {r['comprimento_total']}m"
    c2 = ws["A2"]
    c2.value = kpi
    c2.font  = Font(name="Calibri", size=9)
    c2.fill  = PatternFill("solid", fgColor=C_AZUL_CLR)
    c2.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    row = 4
    mapas = sol["mapas"]
    folhas = sol["folhas"]

    # Totais globais (para totalização ao final)
    tot_global_tam = {t: 0 for t in tamanhos}
    tot_global_folhas = 0

    for mi, mapa in enumerate(mapas):
        npecas = sum(mapa.values())
        comp   = r["comprimento_por_mapa"][mi] if "comprimento_por_mapa" in r else round(npecas * consumo, 4)
        tf_enf = sum(folhas[c][mi] for c in cores)
        ms     = "+".join(f"{mapa[t]}{t}" for t in tamanhos if mapa.get(t, 0) > 0)
        tot_global_folhas += tf_enf

        # Título do enfesto
        ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
        c = ws.cell(row=row, column=1,
                    value=f"ENFESTO {mi+1}  ·  Mapa: {ms}  ·  {npecas} peças × {consumo}m = {comp}m  ·  {tf_enf} folhas")
        c.font      = Font(name="Calibri", size=10, bold=True, color=C_AZUL_ESC)
        c.fill      = PatternFill("solid", fgColor=C_AZUL_CLR)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = _bd()
        ws.row_dimensions[row].height = 20
        row += 1

        # Linha "Peças no mapa"
        _cel(ws, row, 1, "Peças no mapa", negrito=True, fundo=C_CINZA_HDR)
        _cel(ws, row, 2, "—", alinha="center", fundo=C_CINZA_HDR)
        for ti, t in enumerate(tamanhos):
            _cel(ws, row, 3+ti, mapa.get(t, 0), negrito=True, alinha="center", fundo=C_CINZA_HDR)
        _cel(ws, row, 3+len(tamanhos), npecas, negrito=True, alinha="center", fundo=C_CINZA_HDR)
        ws.row_dimensions[row].height = 18
        row += 1

        # Header da tabela
        _cel(ws, row, 1, "Cor", negrito=True, fundo=C_AZUL_MED, cor_txt=C_BRANCO)
        _cel(ws, row, 2, "Folhas", negrito=True, alinha="center", fundo=C_AZUL_MED, cor_txt=C_BRANCO)
        for ti, t in enumerate(tamanhos):
            _cel(ws, row, 3+ti, t, negrito=True, alinha="center", fundo=C_AZUL_MED, cor_txt=C_BRANCO)
        _cel(ws, row, 3+len(tamanhos), "Total", negrito=True, alinha="center", fundo=C_AZUL_MED, cor_txt=C_BRANCO)
        ws.row_dimensions[row].height = 18
        row += 1

        # Linhas por cor
        tot_enf_tam = {t: 0 for t in tamanhos}
        for ci, cor in enumerate(cores):
            f = folhas[cor][mi]
            fundo_l = C_BRANCO if ci % 2 == 0 else C_CINZA_ALT
            _cel(ws, row, 1, cor, negrito=True, fundo=fundo_l)
            _cel(ws, row, 2, f, alinha="center", fundo=fundo_l)
            tot_cor = 0
            for ti, t in enumerate(tamanhos):
                v = f * mapa.get(t, 0)
                _cel(ws, row, 3+ti, v if v > 0 else None, alinha="center", fundo=fundo_l)
                tot_enf_tam[t] += v
                tot_global_tam[t] += v
                tot_cor += v
            _cel(ws, row, 3+len(tamanhos), tot_cor if tot_cor > 0 else None, negrito=True, alinha="center", fundo=fundo_l)
            ws.row_dimensions[row].height = 17
            row += 1

        # Total do enfesto
        _cel(ws, row, 1, f"TOTAL ENFESTO {mi+1}", negrito=True, fundo=C_TOTAL, cor_txt=C_TOTAL_TX)
        _cel(ws, row, 2, tf_enf, negrito=True, alinha="center", fundo=C_TOTAL, cor_txt=C_TOTAL_TX)
        tot_enf_total = 0
        for ti, t in enumerate(tamanhos):
            _cel(ws, row, 3+ti, tot_enf_tam[t] if tot_enf_tam[t] > 0 else None,
                 negrito=True, alinha="center", fundo=C_TOTAL, cor_txt=C_TOTAL_TX)
            tot_enf_total += tot_enf_tam[t]
        _cel(ws, row, 3+len(tamanhos), tot_enf_total, negrito=True, alinha="center", fundo=C_TOTAL, cor_txt=C_TOTAL_TX)
        ws.row_dimensions[row].height = 18
        row += 1

        row = _linha_vazia(ws, row)

    # ── Totalizador geral ──────────────────────────────────────────────────────
    row = _linha_titulo_secao(ws, row, f"TOTALIZADOR GERAL — OPÇÃO {idx}", ncols)
    tot_grade_tam = {t: sum(grade[c].get(t, 0) for c in cores) for t in tamanhos}

    # Linha Grade
    _cel(ws, row, 1, "Grade total", negrito=True, fundo=C_CINZA_HDR)
    _cel(ws, row, 2, "—", alinha="center", fundo=C_CINZA_HDR)
    tot_grade = 0
    for ti, t in enumerate(tamanhos):
        _cel(ws, row, 3+ti, tot_grade_tam[t], alinha="center", fundo=C_CINZA_HDR)
        tot_grade += tot_grade_tam[t]
    _cel(ws, row, 3+len(tamanhos), tot_grade, negrito=True, alinha="center", fundo=C_CINZA_HDR)
    ws.row_dimensions[row].height = 18
    row += 1

    # Linha Corta
    _cel(ws, row, 1, "Corta total", negrito=True, fundo=C_TOTAL, cor_txt=C_TOTAL_TX)
    _cel(ws, row, 2, tot_global_folhas, negrito=True, alinha="center", fundo=C_TOTAL, cor_txt=C_TOTAL_TX)
    tot_corta = 0
    for ti, t in enumerate(tamanhos):
        _cel(ws, row, 3+ti, tot_global_tam[t], negrito=True, alinha="center", fundo=C_TOTAL, cor_txt=C_TOTAL_TX)
        tot_corta += tot_global_tam[t]
    _cel(ws, row, 3+len(tamanhos), tot_corta, negrito=True, alinha="center", fundo=C_TOTAL, cor_txt=C_TOTAL_TX)
    ws.row_dimensions[row].height = 18
    row += 1

    # Linha Diferença
    _cel(ws, row, 1, "Diferença", negrito=True, fundo=C_CINZA_ALT)
    _cel(ws, row, 2, "—", alinha="center", fundo=C_CINZA_ALT)
    tot_diff = 0
    for ti, t in enumerate(tamanhos):
        diff = tot_global_tam[t] - tot_grade_tam[t]
        fundo_d = C_DESVIO0 if diff == 0 else (C_LARANJA if abs(diff) <= 4 else C_VERMELHO)
        cor_d   = C_VERDE_TX if diff == 0 else (C_LARANJA_TX if abs(diff) <= 4 else C_VERM_TX)
        sinal = f"+{diff}" if diff > 0 else str(diff)
        _cel(ws, row, 3+ti, sinal if diff != 0 else "✓", negrito=(diff==0),
             alinha="center", fundo=fundo_d, cor_txt=cor_d)
        tot_diff += diff
    sinal_tot = f"+{tot_diff}" if tot_diff > 0 else str(tot_diff)
    fundo_tot = C_DESVIO0 if tot_diff == 0 else (C_LARANJA if abs(tot_diff) <= 15 else C_VERMELHO)
    cor_tot   = C_VERDE_TX if tot_diff == 0 else (C_LARANJA_TX if abs(tot_diff) <= 15 else C_VERM_TX)
    _cel(ws, row, 3+len(tamanhos), sinal_tot if tot_diff != 0 else "✓", negrito=True,
         alinha="center", fundo=fundo_tot, cor_txt=cor_tot)
    ws.row_dimensions[row].height = 18
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# ABA CONFERÊNCIA POR OPÇÃO
# ══════════════════════════════════════════════════════════════════════════════
def _aba_conf(wb, sol, idx, grade, tamanhos, limites, referencia):
    ws = wb.create_sheet(f"Op{idx}_Conferencia")
    cores = list(grade.keys())
    mapas  = sol["mapas"]
    folhas = sol["folhas"]
    ncols  = 1 + len(tamanhos) * 3 + 2

    ws.column_dimensions["A"].width = 16
    for i in range(2, ncols + 2):
        ws.column_dimensions[get_column_letter(i)].width = 9

    # Cabeçalho
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value = f"CONFERÊNCIA DE GRADE — {referencia.upper()} — OPÇÃO {idx}"
    c.font  = Font(name="Calibri", size=13, bold=True, color=C_BRANCO)
    c.fill  = PatternFill("solid", fgColor=C_AZUL_ESC)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    row = 3
    # Header duplo
    _cel(ws, row, 1, "Cor", negrito=True, fundo=C_AZUL_MED, cor_txt=C_BRANCO)
    col = 2
    for t in tamanhos:
        ws.merge_cells(f"{get_column_letter(col)}{row}:{get_column_letter(col+2)}{row}")
        c = ws.cell(row=row, column=col, value=t)
        c.font      = Font(name="Calibri", size=10, bold=True, color=C_BRANCO)
        c.fill      = PatternFill("solid", fgColor=C_AZUL_MED)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _bd()
        col += 3
    # Colunas total
    ws.merge_cells(f"{get_column_letter(col)}{row}:{get_column_letter(col+1)}{row}")
    c = ws.cell(row=row, column=col, value="Total")
    c.font = Font(name="Calibri", size=10, bold=True, color=C_BRANCO)
    c.fill = PatternFill("solid", fgColor=C_AZUL_ESC)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _bd()
    ws.row_dimensions[row].height = 18
    row += 1

    # Sub-headers
    _cel(ws, row, 1, "", fundo=C_CINZA_HDR)
    col = 2
    for _ in tamanhos:
        for lbl in ["Grade", "Corta", "Ajuste"]:
            _cel(ws, row, col, lbl, negrito=True, alinha="center", fundo=C_CINZA_HDR)
            col += 1
    _cel(ws, row, col, "Grade", negrito=True, alinha="center", fundo=C_CINZA_HDR)
    _cel(ws, row, col+1, "Corta", negrito=True, alinha="center", fundo=C_CINZA_HDR)
    ws.row_dimensions[row].height = 18
    row += 1

    # Linhas por cor
    tot_grade_tam = {t: 0 for t in tamanhos}
    tot_corta_tam = {t: 0 for t in tamanhos}

    for ci, cor in enumerate(cores):
        fundo_l = C_BRANCO if ci % 2 == 0 else C_CINZA_ALT
        _cel(ws, row, 1, cor, negrito=True, fundo=fundo_l)
        col = 2
        tot_g_cor = 0; tot_c_cor = 0
        for t in tamanhos:
            gv = int(grade[cor].get(t, 0))
            ct = sum(folhas[cor][k] * mapas[k].get(t, 0) for k in range(len(mapas)))
            adj = int(ct - gv)
            lo, hi = limites.get(cor, {}).get(t, (-4, 4))
            ok = lo <= adj <= hi

            fundo_adj = C_DESVIO0 if adj == 0 else (C_VERDE if ok else C_VERMELHO)
            cor_adj   = C_VERDE_TX if adj == 0 else (C_VERDE_TX if ok else C_VERM_TX)

            _cel(ws, row, col,   gv,  alinha="center", fundo=fundo_l)
            _cel(ws, row, col+1, ct,  alinha="center", fundo=fundo_l)
            _cel(ws, row, col+2,
                 f"+{adj}" if adj > 0 else ("✓" if adj == 0 else str(adj)),
                 negrito=(adj==0), alinha="center", fundo=fundo_adj, cor_txt=cor_adj)
            col += 3
            tot_grade_tam[t] += gv
            tot_corta_tam[t] += ct
            tot_g_cor += gv; tot_c_cor += ct

        _cel(ws, row, col,   tot_g_cor, negrito=True, alinha="center", fundo=fundo_l)
        _cel(ws, row, col+1, tot_c_cor, negrito=True, alinha="center", fundo=fundo_l)
        ws.row_dimensions[row].height = 17
        row += 1

    # Linha TOTAL
    _cel(ws, row, 1, "TOTAL", negrito=True, fundo=C_TOTAL, cor_txt=C_TOTAL_TX)
    col = 2; tot_g = 0; tot_c = 0
    for t in tamanhos:
        tg = tot_grade_tam[t]; tc = tot_corta_tam[t]; adj = tc - tg
        ok = True  # total pode ter ajuste
        fundo_adj = C_DESVIO0 if adj == 0 else (C_AMARELO if abs(adj) <= 15 else C_VERMELHO)
        cor_adj   = C_VERDE_TX if adj == 0 else (C_AMAR_TX if abs(adj) <= 15 else C_VERM_TX)
        _cel(ws, row, col,   tg, negrito=True, alinha="center", fundo=C_TOTAL, cor_txt=C_TOTAL_TX)
        _cel(ws, row, col+1, tc, negrito=True, alinha="center", fundo=C_TOTAL, cor_txt=C_TOTAL_TX)
        _cel(ws, row, col+2,
             f"+{adj}" if adj > 0 else ("✓" if adj == 0 else str(adj)),
             negrito=True, alinha="center", fundo=fundo_adj, cor_txt=cor_adj)
        col += 3; tot_g += tg; tot_c += tc
    diff_tot = tot_c - tot_g
    _cel(ws, row, col,   tot_g, negrito=True, alinha="center", fundo=C_TOTAL, cor_txt=C_TOTAL_TX)
    fundo_dt = C_DESVIO0 if diff_tot == 0 else (C_AMARELO if abs(diff_tot) <= 15 else C_VERMELHO)
    cor_dt   = C_VERDE_TX if diff_tot == 0 else (C_AMAR_TX if abs(diff_tot) <= 15 else C_VERM_TX)
    _cel(ws, row, col+1,
         f"+{tot_c}" if tot_c != tot_g else str(tot_c),
         negrito=True, alinha="center", fundo=fundo_dt, cor_txt=cor_dt)
    ws.row_dimensions[row].height = 20
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# ABA COMPARATIVO
# ══════════════════════════════════════════════════════════════════════════════
def _aba_comparativo(wb, solucoes, grade, tamanhos, limites, referencia):
    if len(solucoes) < 2:
        return
    ws = wb.create_sheet("Comparativo")
    cores = list(grade.keys())
    n     = len(solucoes)
    ncols = 2 + n * 2 + 1

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 8
    for i in range(3, ncols + 2):
        ws.column_dimensions[get_column_letter(i)].width = 10

    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value = f"COMPARATIVO DE AJUSTES DE GRADE — {referencia.upper()}"
    c.font  = Font(name="Calibri", size=13, bold=True, color=C_BRANCO)
    c.fill  = PatternFill("solid", fgColor=C_AZUL_ESC)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    row = 3
    # Trade-off
    tf = _calcular_tradeoff(solucoes, grade, tamanhos)
    if tf:
        fundo_tf = {
            "ok":    C_VERDE,    "aviso": C_AMARELO, "info":  C_AZUL_CLR
        }.get(tf["nivel"], C_BRANCO)
        cor_tf = {
            "ok":    C_VERDE_TX, "aviso": C_AMAR_TX, "info":  C_AZUL_ESC
        }.get(tf["nivel"], "000000")
        ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
        c = ws.cell(row=row, column=1, value=f"▶  {tf['msg']}")
        c.font      = Font(name="Calibri", size=10, bold=True, color=cor_tf)
        c.fill      = PatternFill("solid", fgColor=fundo_tf)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border    = _bd()
        ws.row_dimensions[row].height = 36
        row += 1

    row = _linha_vazia(ws, row)

    # Resumo rápido
    row = _linha_titulo_secao(ws, row, "RESUMO POR OPÇÃO", ncols)
    _cel(ws, row, 1, "Métrica", negrito=True, fundo=C_CINZA_HDR)
    _cel(ws, row, 2, "", fundo=C_CINZA_HDR)
    for i, s in enumerate(solucoes, 1):
        _cel(ws, row, 2+i, f"Opção {i}", negrito=True, alinha="center",
             fundo=C_AZUL_MED, cor_txt=C_BRANCO)
    ws.row_dimensions[row].height = 18; row += 1

    for lbl, fn in [
        ("Nº de enfestos",        lambda s: s["resumo"]["n_mapas"]),
        ("Média peças/mapa",      lambda s: s["resumo"]["media_pecas_mapa"]),
        ("Total folhas",          lambda s: s["resumo"]["total_folhas"]),
        ("Ajuste de grade (pç)",  lambda s: s["resumo"].get("desvio_total", 0)),
        ("Score otimização",      lambda s: round(s.get("score", 0), 4)),
    ]:
        _cel(ws, row, 1, lbl, fundo=C_CINZA_ALT)
        _cel(ws, row, 2, "", fundo=C_CINZA_ALT)
        for i, s in enumerate(solucoes, 1):
            v = fn(s)
            if lbl.startswith("Ajuste"):
                fundo_v = C_DESVIO0 if v == 0 else (C_AMARELO if v <= 15 else C_VERMELHO)
                cor_v   = C_VERDE_TX if v == 0 else (C_AMAR_TX if v <= 15 else C_VERM_TX)
                _cel(ws, row, 2+i, "✓ Nenhum" if v == 0 else f"{v} peças",
                     negrito=(v==0), alinha="center", fundo=fundo_v, cor_txt=cor_v)
            else:
                _cel(ws, row, 2+i, v, alinha="center", fundo=C_BRANCO)
        ws.row_dimensions[row].height = 17; row += 1

    row = _linha_vazia(ws, row)

    # Tabela detalhe cor/tamanho
    row = _linha_titulo_secao(ws, row, "DETALHE POR COR E TAMANHO", ncols)

    # Header
    _cel(ws, row, 1, "Cor", negrito=True, fundo=C_CINZA_HDR)
    _cel(ws, row, 2, "Tam.", negrito=True, alinha="center", fundo=C_CINZA_HDR)
    _cel(ws, row, 3, "Grade", negrito=True, alinha="center", fundo=C_CINZA_HDR)
    col = 4
    for i, s in enumerate(solucoes, 1):
        ws.merge_cells(f"{get_column_letter(col)}{row}:{get_column_letter(col+1)}{row}")
        c = ws.cell(row=row, column=col, value=f"Opção {i}")
        c.font = Font(name="Calibri", size=10, bold=True, color=C_BRANCO)
        c.fill = PatternFill("solid", fgColor=C_AZUL_MED)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _bd()
        col += 2
    ws.row_dimensions[row].height = 18; row += 1

    # Sub-header Corta / Ajuste
    _cel(ws, row, 1, "", fundo=C_CINZA_HDR)
    _cel(ws, row, 2, "", fundo=C_CINZA_HDR)
    _cel(ws, row, 3, "", fundo=C_CINZA_HDR)
    col = 4
    for _ in solucoes:
        _cel(ws, row, col,   "Corta",  negrito=True, alinha="center", fundo=C_CINZA_HDR)
        _cel(ws, row, col+1, "Ajuste", negrito=True, alinha="center", fundo=C_CINZA_HDR)
        col += 2
    ws.row_dimensions[row].height = 16; row += 1

    # Dados
    tot_grade_t = {t: 0 for t in tamanhos}
    tot_corta_s = [{t: 0 for t in tamanhos} for _ in solucoes]

    for ci, cor in enumerate(cores):
        for ti, t in enumerate(tamanhos):
            gv = int(grade[cor].get(t, 0))
            if gv == 0:
                continue

            vals = []
            for s in solucoes:
                ct = sum(s["folhas"][cor][k] * s["mapas"][k].get(t, 0)
                         for k in range(len(s["mapas"])))
                vals.append((ct, ct - gv))

            min_adj = min(abs(v[1]) for v in vals)
            tem_dif = len(set(v[1] for v in vals)) > 1
            fundo_l = C_AMARELO if tem_dif else (C_BRANCO if ci % 2 == 0 else C_CINZA_ALT)

            _cel(ws, row, 1, cor, fundo=fundo_l)
            _cel(ws, row, 2, t, alinha="center", fundo=fundo_l)
            _cel(ws, row, 3, gv, alinha="center", fundo=fundo_l)
            col = 4
            for vi, (ct, adj) in enumerate(vals):
                lo, hi = limites.get(cor, {}).get(t, (-4, 4))
                ok = lo <= adj <= hi
                fundo_adj = C_DESVIO0 if adj == 0 else (C_VERDE if ok and abs(adj) == min_adj else C_VERMELHO if not ok else C_BRANCO)
                cor_adj   = C_VERDE_TX if adj == 0 else (C_VERDE_TX if ok and abs(adj) == min_adj else C_VERM_TX if not ok else "000000")
                _cel(ws, row, col, ct, alinha="center", fundo=fundo_l)
                _cel(ws, row, col+1,
                     "✓" if adj == 0 else (f"+{adj}" if adj > 0 else str(adj)),
                     negrito=(adj==0), alinha="center", fundo=fundo_adj, cor_txt=cor_adj)
                col += 2
                tot_corta_s[vi][t] += ct
            tot_grade_t[t] += gv
            ws.row_dimensions[row].height = 16; row += 1

    row = _linha_vazia(ws, row)

    # Linha totais
    _cel(ws, row, 1, "TOTAL", negrito=True, fundo=C_TOTAL, cor_txt=C_TOTAL_TX)
    _cel(ws, row, 2, "", fundo=C_TOTAL)
    tot_g = sum(tot_grade_t.values())
    _cel(ws, row, 3, tot_g, negrito=True, alinha="center", fundo=C_TOTAL, cor_txt=C_TOTAL_TX)
    col = 4
    for vi, tc_dict in enumerate(tot_corta_s):
        tc = sum(tc_dict.values())
        adj = tc - tot_g
        fundo_v = C_DESVIO0 if adj == 0 else (C_AMARELO if abs(adj) <= 15 else C_VERMELHO)
        cor_v   = C_VERDE_TX if adj == 0 else (C_AMAR_TX if abs(adj) <= 15 else C_VERM_TX)
        _cel(ws, row, col, tc, negrito=True, alinha="center", fundo=C_TOTAL, cor_txt=C_TOTAL_TX)
        _cel(ws, row, col+1,
             "✓" if adj == 0 else (f"+{adj}" if adj > 0 else str(adj)),
             negrito=True, alinha="center", fundo=fundo_v, cor_txt=cor_v)
        col += 2
    ws.row_dimensions[row].height = 20

    # Legenda
    row += 2
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1,
                value="Legenda:  ✓ = exato  |  Verde = menor desvio entre opções  |  Vermelho = fora da tolerância  |  Amarelo = linha com diferença entre as opções")
    c.font      = Font(name="Calibri", size=8, italic=True, color="666666")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 16
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# PONTO DE ENTRADA
# ══════════════════════════════════════════════════════════════════════════════
def exportar(solucoes, grade, tamanhos, limites, config, referencia, pasta_saida="dados/resultados"):
    if not EXCEL_OK:
        raise ImportError("openpyxl não instalado. Execute: pip install openpyxl")

    os.makedirs(pasta_saida, exist_ok=True)
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome = referencia.replace(" ", "_").replace("/", "-")[:40]
    cam  = os.path.join(pasta_saida, f"plano_corte_{nome}_{ts}.xlsx")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remover aba vazia padrão

    consumo = float(config.get("consumo_peca_m", 1.0645))
    for s in solucoes:
        s["consumo"] = consumo

    # 1. Resumo
    _aba_resumo(wb, solucoes, grade, tamanhos, referencia, config)

    # 2. Plano + Conferência por opção
    for idx, sol in enumerate(solucoes, 1):
        _aba_plano(wb, sol, idx, grade, tamanhos, referencia, config)
        _aba_conf(wb, sol, idx, grade, tamanhos, limites, referencia)

    # 3. Comparativo (se > 1 opção)
    if len(solucoes) > 1:
        _aba_comparativo(wb, solucoes, grade, tamanhos, limites, referencia)

    wb.save(cam)
    return cam


# ══════════════════════════════════════════════════════════════════════════════
# EXPORTAR MULTI-REF
# ══════════════════════════════════════════════════════════════════════════════
def _aba_resumo_multiref(wb, solucoes, referencia, config=None):
    config = config or {}
    """Aba de resumo comparando opções multi-ref."""
    ws = wb.create_sheet("Resumo")
    n = len(solucoes)
    ncols = max(n + 1, 3)
    ws.column_dimensions["A"].width = 34
    for i in range(2, ncols + 2):
        ws.column_dimensions[get_column_letter(i)].width = 22

    ts = datetime.now().strftime("%d/%m/%Y %H:%M")

    ws.merge_cells(f"A1:{get_column_letter(ncols + 1)}1")
    c = ws["A1"]
    c.value = f"PCP ENFESTOS COMBINADO — {referencia.upper()}"
    c.font  = Font(name="Calibri", size=14, bold=True, color=C_BRANCO)
    c.fill  = PatternFill("solid", fgColor=C_AZUL_ESC)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{get_column_letter(ncols + 1)}2")
    c2 = ws["A2"]
    n_refs = len(solucoes[0]["refs_sol"]) if solucoes else 0
    c2.value = (f"Gerado em {ts}  |  Enfesto combinado: {n_refs} refs  |  {n} opcao(oes)"
                f"  |  " + _resumo_parametros_txt(config))
    c2.font  = Font(name="Calibri", size=9, color="444444")
    c2.fill  = PatternFill("solid", fgColor=C_AZUL_CLR)
    c2.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    row = 4
    row = _linha_titulo_secao(ws, row, "COMPARATIVO DE OPÇÕES", ncols + 1)

    _cel(ws, row, 1, "Métrica", negrito=True, fundo=C_CINZA_HDR)
    for i in range(n):
        _cel(ws, row, i + 2, f"Opção {i+1}", negrito=True, fundo=C_AZUL_MED,
             cor_txt=C_BRANCO, alinha="center")
    ws.row_dimensions[row].height = 18
    row += 1

    metricas = [
        ("Nº de enfestos combinados", [s["n_mapas"] for s in solucoes]),
        ("Total de folhas",            [s["resumo"]["total_folhas"] for s in solucoes]),
        ("Desvio total (peças)",       [s["resumo"]["desvio_total"] for s in solucoes]),
        ("Comprimento total (m)",      [s["resumo"]["comprimento_total"] for s in solucoes]),
        ("Média peças/mapa",           [s["resumo"]["media_pecas_mapa"] for s in solucoes]),
    ]
    for label, vals in metricas:
        _cel(ws, row, 1, label, fundo=C_CINZA_ALT)
        for i, v in enumerate(vals):
            _cel(ws, row, i + 2, v, alinha="center", fundo=C_BRANCO)
        ws.row_dimensions[row].height = 18
        row += 1

    row = _linha_vazia(ws, row)
    row = _linha_titulo_secao(ws, row, "COMPRIMENTOS POR ENFESTO", ncols + 1)

    for oi, sol in enumerate(solucoes):
        _cel(ws, row, 1, f"Opção {oi+1}:", negrito=True, fundo=C_CINZA_HDR)
        ws.row_dimensions[row].height = 16
        row += 1
        for ki, comp in enumerate(sol["comprimentos"]):
            refs_desc = "  |  ".join(
                "+".join(f"{sol['refs_sol'][ri]['mapas'][ki].get(t, 0)}{t}"
                         for t in (list(sol['refs_sol'][ri]['mapas'][ki].keys()) if ki < len(sol['refs_sol'][ri]['mapas']) else [])
                         if sol['refs_sol'][ri]['mapas'][ki].get(t, 0) > 0)
                for ri in range(len(sol["refs_sol"]))
            )
            ws.merge_cells(f"A{row}:{get_column_letter(ncols + 1)}{row}")
            c = ws.cell(row=row, column=1,
                        value=f"  Enfesto {ki+1}: {comp}m  [{refs_desc}]")
            c.font = Font(name="Calibri", size=9)
            c.fill = PatternFill("solid", fgColor=C_BRANCO if ki % 2 == 0 else C_CINZA_ALT)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = _bd()
            ws.row_dimensions[row].height = 16
            row += 1
        row = _linha_vazia(ws, row)

    return ws


def _aba_plano_multiref(wb, sol, idx, tamanhos, referencia):
    """Aba de plano por opção — mostra cada ref separadamente."""
    ws = wb.create_sheet(f"Op{idx}_Plano")
    n_refs  = len(sol["refs_sol"])
    n_mapas = sol["n_mapas"]
    ncols   = 2 + len(tamanhos) + 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 10
    for i in range(3, 3 + len(tamanhos)):
        ws.column_dimensions[get_column_letter(i)].width = 9
    ws.column_dimensions[get_column_letter(3 + len(tamanhos))].width = 10

    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value = f"PLANO COMBINADO — {referencia.upper()} — OPÇÃO {idx}"
    c.font  = Font(name="Calibri", size=13, bold=True, color=C_BRANCO)
    c.fill  = PatternFill("solid", fgColor=C_AZUL_ESC)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    r = sol["resumo"]
    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    dev = r.get("desvio_total", 0)
    kpi = (f"Enfestos: {r['n_mapas']}  |  Total folhas: {r['total_folhas']}  |  "
           f"Ajuste total: {'Sem ajuste ✓' if dev == 0 else str(dev)+' peças'}  |  "
           f"Comp. total: {r['comprimento_total']}m")
    c2 = ws["A2"]
    c2.value = kpi
    c2.font  = Font(name="Calibri", size=9)
    c2.fill  = PatternFill("solid", fgColor=C_AZUL_CLR)
    c2.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    row = 4

    for ki in range(n_mapas):
        comp = sol["comprimentos"][ki] if ki < len(sol["comprimentos"]) else "—"

        # Título do enfesto combinado
        ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
        c = ws.cell(row=row, column=1,
                    value=f"ENFESTO COMBINADO {ki+1}  ·  {comp}m  ·  {n_refs} refs")
        c.font      = Font(name="Calibri", size=11, bold=True, color=C_BRANCO)
        c.fill      = PatternFill("solid", fgColor=C_AZUL_ESC)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = _bd()
        ws.row_dimensions[row].height = 22
        row += 1

        # Por ref
        for ri, ref in enumerate(sol["refs_sol"]):
            mapa   = ref["mapas"][ki] if ki < len(ref["mapas"]) else {}
            folhas = ref["folhas"]
            cores  = list(ref["grade"].keys())
            consumo = float(ref.get("consumo", 1.0645))
            npecas  = sum(mapa.values())
            comp_ref = round(npecas * consumo, 4)
            tf_ref   = sum(folhas[c][ki] for c in cores)
            ms = "+".join(f"{mapa[t]}{t}" for t in tamanhos if mapa.get(t, 0) > 0)

            ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
            c = ws.cell(row=row, column=1,
                        value=f"  ↳ {ref.get('nome','Ref '+str(ri+1))}  ·  Mapa: {ms}  ·  {npecas}pç × {consumo}m = {comp_ref}m  ·  {tf_ref} folhas")
            c.font      = Font(name="Calibri", size=9, bold=True, color=C_AZUL_ESC)
            c.fill      = PatternFill("solid", fgColor=C_AZUL_CLR)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border    = _bd()
            ws.row_dimensions[row].height = 18
            row += 1

            # Header
            _cel(ws, row, 1, "Cor", negrito=True, fundo=C_CINZA_HDR)
            _cel(ws, row, 2, "Folhas", negrito=True, alinha="center", fundo=C_CINZA_HDR)
            for ti, t in enumerate(tamanhos):
                _cel(ws, row, 3+ti, t, negrito=True, alinha="center", fundo=C_CINZA_HDR)
            _cel(ws, row, 3+len(tamanhos), "Total", negrito=True, alinha="center", fundo=C_CINZA_HDR)
            ws.row_dimensions[row].height = 16
            row += 1

            for ci, cor in enumerate(cores):
                f = folhas[cor][ki] if ki < len(folhas.get(cor, [])) else 0
                fundo_l = C_BRANCO if ci % 2 == 0 else C_CINZA_ALT
                _cel(ws, row, 1, cor, fundo=fundo_l)
                _cel(ws, row, 2, f, alinha="center", fundo=fundo_l)
                tot_cor = 0
                for ti, t in enumerate(tamanhos):
                    v = f * mapa.get(t, 0)
                    _cel(ws, row, 3+ti, v if v > 0 else None, alinha="center", fundo=fundo_l)
                    tot_cor += v
                _cel(ws, row, 3+len(tamanhos), tot_cor if tot_cor > 0 else None,
                     negrito=True, alinha="center", fundo=fundo_l)
                ws.row_dimensions[row].height = 16
                row += 1

            _cel(ws, row, 1, f"TOTAL {ref.get('nome','Ref '+str(ri+1))}", negrito=True,
                 fundo=C_TOTAL, cor_txt=C_TOTAL_TX)
            _cel(ws, row, 2, tf_ref, negrito=True, alinha="center", fundo=C_TOTAL, cor_txt=C_TOTAL_TX)
            tot_t = {t: sum((folhas[c][ki] if ki < len(folhas.get(c, [])) else 0) * mapa.get(t, 0) for c in cores)
                     for t in tamanhos}
            for ti, t in enumerate(tamanhos):
                _cel(ws, row, 3+ti, tot_t[t] if tot_t[t] > 0 else None,
                     negrito=True, alinha="center", fundo=C_TOTAL, cor_txt=C_TOTAL_TX)
            _cel(ws, row, 3+len(tamanhos), sum(tot_t.values()),
                 negrito=True, alinha="center", fundo=C_TOTAL, cor_txt=C_TOTAL_TX)
            ws.row_dimensions[row].height = 17
            row += 1
            row = _linha_vazia(ws, row)

    return ws


def _aba_conf_multiref(wb, sol, idx, tamanhos, referencia):
    """Aba conferência multi-ref — grade vs cortado por ref."""
    ws = wb.create_sheet(f"Op{idx}_Conferencia")
    ncols = 2 + len(tamanhos) * 3 + 2

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16
    for i in range(3, ncols + 2):
        ws.column_dimensions[get_column_letter(i)].width = 9

    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value = f"CONFERÊNCIA COMBINADA — {referencia.upper()} — OPÇÃO {idx}"
    c.font  = Font(name="Calibri", size=13, bold=True, color=C_BRANCO)
    c.fill  = PatternFill("solid", fgColor=C_AZUL_ESC)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    row = 3
    for ri, ref in enumerate(sol["refs_sol"]):
        row = _linha_titulo_secao(ws, row, f"{ref.get('nome', 'Ref '+str(ri+1))}  (consumo: {ref.get('consumo', 1.0645)}m/pç)", ncols)
        mapas  = ref["mapas"]
        folhas = ref["folhas"]
        limites = ref.get("limites", {})
        grade  = ref["grade"]
        cores  = list(grade.keys())
        n_mapas = len(mapas)

        # Headers
        _cel(ws, row, 1, "Cor", negrito=True, fundo=C_AZUL_MED, cor_txt=C_BRANCO)
        _cel(ws, row, 2, "", fundo=C_AZUL_MED)
        col = 3
        for t in tamanhos:
            ws.merge_cells(f"{get_column_letter(col)}{row}:{get_column_letter(col+2)}{row}")
            c = ws.cell(row=row, column=col, value=t)
            c.font = Font(name="Calibri", size=10, bold=True, color=C_BRANCO)
            c.fill = PatternFill("solid", fgColor=C_AZUL_MED)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _bd()
            col += 3
        ws.merge_cells(f"{get_column_letter(col)}{row}:{get_column_letter(col+1)}{row}")
        c = ws.cell(row=row, column=col, value="Total")
        c.font = Font(name="Calibri", size=10, bold=True, color=C_BRANCO)
        c.fill = PatternFill("solid", fgColor=C_AZUL_ESC)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _bd()
        ws.row_dimensions[row].height = 18
        row += 1

        _cel(ws, row, 1, "", fundo=C_CINZA_HDR)
        _cel(ws, row, 2, "", fundo=C_CINZA_HDR)
        col = 3
        for _ in tamanhos:
            for lbl in ["Grade", "Corta", "Ajuste"]:
                _cel(ws, row, col, lbl, negrito=True, alinha="center", fundo=C_CINZA_HDR)
                col += 1
        _cel(ws, row, col, "Grade", negrito=True, alinha="center", fundo=C_CINZA_HDR)
        _cel(ws, row, col+1, "Corta", negrito=True, alinha="center", fundo=C_CINZA_HDR)
        ws.row_dimensions[row].height = 16
        row += 1

        tot_grade_tam = {t: 0 for t in tamanhos}
        tot_corta_tam = {t: 0 for t in tamanhos}

        for ci, cor in enumerate(cores):
            fundo_l = C_BRANCO if ci % 2 == 0 else C_CINZA_ALT
            _cel(ws, row, 1, cor, negrito=True, fundo=fundo_l)
            _cel(ws, row, 2, "", fundo=fundo_l)
            col = 3
            tot_g_cor = 0; tot_c_cor = 0
            for t in tamanhos:
                gv = int(grade[cor].get(t, 0))
                ct = sum(folhas[cor][k] * mapas[k].get(t, 0)
                         for k in range(n_mapas) if k < len(folhas.get(cor, [])))
                adj = int(ct - gv)
                lo, hi = limites.get(cor, {}).get(t, (-4, 4))
                ok = lo <= adj <= hi
                fundo_adj = C_DESVIO0 if adj == 0 else (C_VERDE if ok else C_VERMELHO)
                cor_adj   = C_VERDE_TX if adj == 0 else (C_VERDE_TX if ok else C_VERM_TX)
                _cel(ws, row, col,   gv,  alinha="center", fundo=fundo_l)
                _cel(ws, row, col+1, ct,  alinha="center", fundo=fundo_l)
                _cel(ws, row, col+2,
                     f"+{adj}" if adj > 0 else ("✓" if adj == 0 else str(adj)),
                     negrito=(adj == 0), alinha="center", fundo=fundo_adj, cor_txt=cor_adj)
                col += 3
                tot_grade_tam[t] += gv; tot_corta_tam[t] += ct
                tot_g_cor += gv; tot_c_cor += ct
            _cel(ws, row, col,   tot_g_cor, negrito=True, alinha="center", fundo=fundo_l)
            _cel(ws, row, col+1, tot_c_cor, negrito=True, alinha="center", fundo=fundo_l)
            ws.row_dimensions[row].height = 17
            row += 1

        _cel(ws, row, 1, "TOTAL", negrito=True, fundo=C_TOTAL, cor_txt=C_TOTAL_TX)
        _cel(ws, row, 2, "", fundo=C_TOTAL)
        col = 3; tot_g = 0; tot_c = 0
        for t in tamanhos:
            tg = tot_grade_tam[t]; tc = tot_corta_tam[t]; adj = tc - tg
            fundo_adj = C_DESVIO0 if adj == 0 else (C_AMARELO if abs(adj) <= 15 else C_VERMELHO)
            cor_adj   = C_VERDE_TX if adj == 0 else (C_AMAR_TX if abs(adj) <= 15 else C_VERM_TX)
            _cel(ws, row, col,   tg, negrito=True, alinha="center", fundo=C_TOTAL, cor_txt=C_TOTAL_TX)
            _cel(ws, row, col+1, tc, negrito=True, alinha="center", fundo=C_TOTAL, cor_txt=C_TOTAL_TX)
            _cel(ws, row, col+2,
                 f"+{adj}" if adj > 0 else ("✓" if adj == 0 else str(adj)),
                 negrito=True, alinha="center", fundo=fundo_adj, cor_txt=cor_adj)
            col += 3; tot_g += tg; tot_c += tc
        diff_tot = tot_c - tot_g
        _cel(ws, row, col,   tot_g, negrito=True, alinha="center", fundo=C_TOTAL, cor_txt=C_TOTAL_TX)
        fundo_dt = C_DESVIO0 if diff_tot == 0 else (C_AMARELO if abs(diff_tot) <= 15 else C_VERMELHO)
        cor_dt   = C_VERDE_TX if diff_tot == 0 else (C_AMAR_TX if abs(diff_tot) <= 15 else C_VERM_TX)
        _cel(ws, row, col+1,
             f"+{tot_c}" if tot_c != tot_g else str(tot_c),
             negrito=True, alinha="center", fundo=fundo_dt, cor_txt=cor_dt)
        ws.row_dimensions[row].height = 20
        row += 1
        row = _linha_vazia(ws, row)

    return ws


def exportar_multiref(solucoes, tamanhos, referencia, config, pasta_saida="dados/resultados"):
    """Exporta resultados multi-ref combinado para Excel."""
    if not EXCEL_OK:
        raise ImportError("openpyxl não instalado. Execute: pip install openpyxl")

    os.makedirs(pasta_saida, exist_ok=True)
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome = referencia.replace(" ", "_").replace("/", "-")[:40]
    cam  = os.path.join(pasta_saida, f"plano_combinado_{nome}_{ts}.xlsx")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _aba_resumo_multiref(wb, solucoes, referencia, config)

    for idx, sol in enumerate(solucoes, 1):
        _aba_plano_multiref(wb, sol, idx, tamanhos, referencia)
        _aba_conf_multiref(wb, sol, idx, tamanhos, referencia)

    wb.save(cam)
    return cam


# ── Exportação de Alocação de Rolos ──────────────────────────────────────────

def _aba_resumo_alocacao(wb, resultado, referencia):
    """Aba de resumo geral da alocacao de rolos."""
    ws = wb.create_sheet("Resumo Alocacao")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20

    r = 1
    _cel(ws, r, 1, f"Alocacao de Rolos — {referencia}", negrito=True, tamanho=13,
         fundo=C_AZUL_ESC, cor_txt=C_BRANCO)
    ws.merge_cells(f"A{r}:B{r}")
    r += 1

    resumo = resultado.get("resumo_geral", {})
    campos = [
        ("Tecido usado total (m)",       resumo.get("tecido_usado_total_m", 0)),
        ("Ponta reaproveitavel total (m)",resumo.get("ponta_estoque_total_m", 0)),
        ("Refugo real total (m)",         resumo.get("refugo_real_total_m", 0)),
        ("Refugo (% sobre nominal)",      resumo.get("refugo_percentual_medio", 0)),
        ("Total de sub-enfestos",         resumo.get("n_sub_enfestos_total", 0)),
    ]
    for label, valor in campos:
        _cel(ws, r, 1, label, fundo=C_CINZA_HDR, negrito=True)
        _cel(ws, r, 2, valor, alinha="right")
        r += 1

    cores_deficit = resumo.get("cores_com_deficit", [])
    if cores_deficit:
        r += 1
        _cel(ws, r, 1, "Cores com deficit:", negrito=True, fundo=C_VERMELHO, cor_txt=C_VERM_TX)
        _cel(ws, r, 2, ", ".join(cores_deficit), fundo=C_VERMELHO, cor_txt=C_VERM_TX)
        r += 1

    alertas = resumo.get("alertas", [])
    if alertas:
        r += 1
        _cel(ws, r, 1, "Alertas", negrito=True, fundo=C_AZUL_MED, cor_txt=C_BRANCO)
        r += 1
        for alerta in alertas:
            _cel(ws, r, 1, alerta, fundo=C_AMARELO, cor_txt=C_AMAR_TX)
            ws.merge_cells(f"A{r}:B{r}")
            r += 1


def _aba_cor_alocacao(wb, cor, cor_res):
    """Aba de detalhe por cor: rolos, sub-enfestos, pontas."""
    nome_aba = f"Rolos {cor[:20]}"
    ws = wb.create_sheet(nome_aba)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    r = 1
    _cel(ws, r, 1, f"Cor: {cor}", negrito=True, tamanho=12,
         fundo=C_AZUL_ESC, cor_txt=C_BRANCO)
    ws.merge_cells(f"A{r}:G{r}")
    r += 2

    # KPIs da cor
    kpis = [
        ("Tecido usado (m)",         cor_res.get("tecido_usado_m", 0)),
        ("Ponta estoque (m)",         cor_res.get("ponta_estoque_total_m", 0)),
        ("Refugo real (m)",           cor_res.get("refugo_real_m", 0)),
        ("Refugo (%)",                cor_res.get("refugo_percentual", 0)),
        ("Sub-enfestos",              cor_res.get("n_sub_enfestos", 0)),
        ("Tecido a comprar (m)",      cor_res.get("tecido_a_comprar_m", 0)),
    ]
    for label, valor in kpis:
        fundo = C_VERMELHO if label == "Tecido a comprar (m)" and valor > 0 else C_CINZA_HDR
        _cel(ws, r, 1, label, negrito=True, fundo=fundo)
        _cel(ws, r, 2, valor, alinha="right")
        r += 1

    r += 1

    # Tabela de rolos
    headers = ["Rolo", "Nominal (m)", "Seguro (m)", "Usado (m)", "Ponta (m)", "Classe", "Sub-enfestos"]
    for col, h in enumerate(headers, 1):
        _cel(ws, r, col, h, negrito=True, fundo=C_AZUL_MED, cor_txt=C_BRANCO, alinha="center")
    r += 1

    for rolo in cor_res.get("rolos", []):
        n_sub  = len(rolo.get("sub_enfestos", []))
        fundo  = C_VERDE if rolo["ponta_classe"] == "estoque" else C_CINZA_ALT
        vals   = [
            rolo["indice"] + 1,
            rolo["comprimento_nominal_m"],
            rolo["comprimento_seguro_m"],
            rolo["usado_m"],
            rolo["ponta_m"],
            rolo["ponta_classe"],
            n_sub,
        ]
        for col, v in enumerate(vals, 1):
            _cel(ws, r, col, v, fundo=fundo, alinha="center" if col in (1, 6, 7) else "right")
        r += 1

        # Sub-enfestos deste rolo
        for sub in rolo.get("sub_enfestos", []):
            _cel(ws, r, 2, f"  Mapa {sub['mapa_id']}", fundo=C_CINZA_ALT)
            _cel(ws, r, 3, f"{sub['n_camadas']} camadas", fundo=C_CINZA_ALT, alinha="right")
            _cel(ws, r, 4, sub["comp_camada"], fundo=C_CINZA_ALT, alinha="right")
            _cel(ws, r, 5, sub["comp_total"], fundo=C_CINZA_ALT, alinha="right")
            _cel(ws, r, 6, f"(+{sub['margem_m']}m faca)", fundo=C_CINZA_ALT)
            r += 1


def exportar_alocacao(resultado, referencia, pasta_saida):
    """
    Exporta resultado do alocador de rolos para Excel.

    Args:
        resultado: saida de alocar_rolos()
        referencia: nome da referencia/pedido
        pasta_saida: diretorio onde salvar

    Returns:
        str: caminho do arquivo gerado
    """
    if not EXCEL_OK:
        raise ImportError("openpyxl nao instalado.")

    os.makedirs(pasta_saida, exist_ok=True)
    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    ref = referencia.replace(" ", "_")[:30]
    cam = os.path.join(pasta_saida, f"alocacao_rolos_{ref}_{ts}.xlsx")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _aba_resumo_alocacao(wb, resultado, referencia)

    for cor, cor_res in sorted(resultado.get("por_cor", {}).items()):
        _aba_cor_alocacao(wb, cor, cor_res)

    wb.save(cam)
    return cam
